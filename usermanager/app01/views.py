from django.http import HttpResponse
from django.shortcuts import render,redirect
from app01.models import Department,UserInfo,Number,admin,Boss,City
from django import forms
from app01 import models
from django.core.validators import RegexValidator,ValidationError
from app01.utils.bootstrap import bootstrapModelForm
from app01.utils.encrypt import md5
from app01.utils.code import check_code
from django.http import JsonResponse
import random
import os
from datetime import datetime
from django.conf import settings
# Create your views here.
def depart_show(request):
    #获取部门表中的数据
    department = Department.objects.all()
    return render(request,"depart_list.html",{'department':department})

def depart_add(request):
    #新增部门数据
    if request.method =="GET":
        return render(request,"depart_add.html")
    title = request.POST.get("title")
    Department.objects.create(title=title)
    return redirect("/depart/list/")

def depart_delete(request):
    #删除部门
    #通过uid找到当前部门信息
    uid = request.GET.get('uid')
    #删除
    Department.objects.filter(id=uid).delete()
    #跳转回列表
    return redirect("/depart/list/")

def depart_edit(request,uid):
    if request.method == "GET":
        departmentById = Department.objects.filter(id=uid).first()
        return render(request,"depart_edit.html",{'departmentById':departmentById})
    #根据id找到当前数据库中的数据
    title_Edit = request.POST.get("title")
    Department.objects.filter(id=uid).update(title=title_Edit)
    return redirect("/depart/list/")


from openpyxl import load_workbook
#批量上传excel文件
def depart_multi(request):

    # 1.获取文件的对象
    exc_obj = request.FILES.get("exc")

    # 2.对象传递给openpyxl,由openpyxl读取文件内容
    wb = load_workbook(exc_obj)
    sheet = wb.worksheets[0] #excel的第一个sheet

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):  # 最小行2，从第2行开始
        text = row[0].value
        #print(text)
        #判断
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect("/depart/list/")

def user_show(request):
    userinfo = UserInfo.objects.all()
    #userinfo.gender ## 1/2
    #userinfo.get_字段名称_display()
    #userinfo.depart_id 拿原始的值，
    #userinfo.depart.title   拿值对应的信息 （有表关联情况）
    #
    return render(request,"user_list.html",{'userinfo':userinfo})


#使用modelform实现用户添加
class UserModelFrom(bootstrapModelForm):
    #约束name字段最小长度为3
    name = forms.CharField(min_length=3,label="用户名")
    #编写正则表达式
    #password = forms.CharField(label="密码",validators=[RegexValidator(r'^159[0-9]+$','数字必须以159开头’)])
    class Meta:
        model = models.UserInfo
        fields = ["name", "password", "age", "account", "create_time", "gender", "depart"]
        widgts = {
            "create_time": forms.DateTimeInput
        }


def user_add(request):
    title = "新建用户"
    if request.method=="GET":
        form = UserModelFrom()
        return render(request,"user_add.html",{"form":form,"title": title })
    #用post提交数据，数据校验
    form = UserModelFrom(data=request.POST)
    if form.is_valid():
        #校验成功，保存到数据库
        form.save()
        return redirect('/user/list/')
    else:
        return render(request,'change.html',{'form':form})

#用户编辑
def user_edit(request,uid):
    userObj = UserInfo.objects.filter(id=uid).first()
    if request.method == "GET":
        #根据id找到当前数据库中的数据
       # userObj = UserInfo.objects.filter(id=uid).first()
        form = UserModelFrom(instance=userObj) # 把默认数据传递进去
        return render(request, "user_edit.html", {'form': form})
    #用post提交数据，数据校验
    # 数据修改的信息，给数据库的哪一行做修改？
    #userObj = UserInfo.objects.filter(id=uid).first()
    form = UserModelFrom(instance=userObj,data=request.POST)
    if form.is_valid():
        #校验成功，保存到数据库
       # form.instance.password = 999    除了用户输入的值
        form.save()
        return redirect('/user/list/')
    else:
        return render(request,'user_add.html',{'form':form})

#用户删除
def user_delete(request,uid):
    #删除部门
    #删除指定id的用户信息
    UserInfo.objects.filter(id=uid).delete()
    #跳转回列表
    return redirect("/user/list/")

#靓号列表
def pretty_list(request):
    ##靓号列表 搜索功能
    data_dict = {}
    search_data = request.GET.get('q','') #有参数就拿到搜索的数据，无就为空
    if search_data:
        data_dict["mobile__contains"] = search_data
        #默认选择全部，筛选条件为 **data__dict，包含search_data中的值
    number = Number.objects.filter(**data_dict).order_by("-level")#按级别排序

    return render(request, "pretty_list.html", {'number': number,"search_data": search_data})

#使用modelform实现靓号添加
class PrettyModelFrom(bootstrapModelForm):
    #正则表达式校验
    mobile = forms.CharField(label="手机号", validators=[RegexValidator(r'^159[0-9]+$', "格式错误，要以159开头")])
    class Meta:
        model = models.Number
        fields = ["mobile", "price", "level", "status"]
        #fields = "__all__"
        #exclude= ['level'] 排除level字段
    # def __init__(self,*args,**kwargs):
    #     super().__init__(*args,**kwargs)
    #     # 循环找到所有插件，添加class
    #     for name,field in self.fields.items():
    #         if field.widget.attrs :
    #             field.widget.attrs["class"] = "form-control"
    #         else:
    #             field.widget.attrs = {
    #                 "class": "form-control"
    #             }
    #方式2 验证
    def clean_mobile(self):
        txt_mobile = self.cleaned_data["mobile"]
        exists = Number.objects.filter(mobile=txt_mobile)
        if exists:
            raise ValidationError("手机号已存在")
        #验证通过，返回输入的值
        return txt_mobile

def pretty_add(request):
    title = "添加靓号"
    if request.method == "GET":
        form = PrettyModelFrom()
        return render(request, "pretty_add.html", {"form": form,'title': title})
    # 用post提交数据，数据校验
    form = PrettyModelFrom(data=request.POST)
    if form.is_valid():
        # 校验成功，保存到数据库
        form.save()
        return redirect('/pretty/list/')
    return render(request, 'change.html', {'form': form})

#使用modelform实现靓号编辑

class PrettyEditModelFrom(bootstrapModelForm):
    mobile = forms.CharField(label="手机号", validators=[RegexValidator(r'^159[0-9]+$', "格式错误，要以159开头")])
    #编辑页面手机号不可编辑
    #mobile = forms.CharField(disabled=True,label="手机号")
    class Meta:
        model = models.Number
        fields = ["mobile","price", "level", "status"]
        #fields = "__all__"
        #exclude= ['level'] 排除level字段
    # def __init__(self,*args,**kwargs):
    #     super().__init__(*args,**kwargs)
    #     # 循环找到所有插件，添加class
    #     for name,field in self.fields.items():
    #         if field.widget.attrs :
    #             field.widget.attrs["class"] = "form-control"
    #         else:
    #             field.widget.attrs = {
    #                 "class": "form-control"
    #             }

    # 方式2 验证
    def clean_mobile(self):
        #当前编辑的一行的ID
        #print(self.instance.pk)
        txt_mobile = self.cleaned_data["mobile"]
        # 当然，编辑页面手机号不可编辑时就无效 排除自己那列，如果已存在，，，
        exists = Number.objects.filter(mobile=txt_mobile).exclude(id=self.instance.pk).exists()
        if exists:
            raise ValidationError("手机号已存在")
        #验证通过，返回输入的值
        return txt_mobile

def pretty_edit(request,uid):
    numObj = Number.objects.filter(id=uid).first()

    if request.method == "GET":
        form = PrettyEditModelFrom(instance=numObj)
        return render(request,"pretty_add.html",{'form':form})

    form = PrettyEditModelFrom(instance=numObj,data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/pretty/list')
    return render(request,"pretty_add.html",{'form':form})

def pretty_delete(request,uid):
    Number.objects.filter(id=uid).delete()
    return redirect("/pretty/list")


def admin_list(request):
    ##管理员列表
    #检测用户是否登录，已登录，继续，未登录，跳转登录页
    #用户发来的请求， 获取cookie随机字符串，拿字符串看看session中有没有
    # info = request.session.get["info"] #到这一步一定拿到session了
    # if not info:
    #     return redirect("/login/")

    ##靓号列表 搜索功能
    data_dict = {}
    search_data = request.GET.get('q','') #有参数就拿到搜索的数据，无就为空
    if search_data:
        data_dict["username__contains"] = search_data
        #默认选择全部，筛选条件为 **data__dict，包含search_data中的值
    queryset = admin.objects.filter(**data_dict)
    context = {
        'queryset': queryset,
        'search_data': search_data,

    }
    return render(request,"admin_list.html",context)


#管理员添加的modelform
class AdminModelForm(bootstrapModelForm):
    #实现二次验证密码
    confirm_password = forms.CharField(
        label="确认密码",
        widget = forms.PasswordInput(render_value=True)
    )
    class Meta:
        model = models.admin
        fields = ["username","password","confirm_password"]
        #定制输入框类型
        widgets = {
            "password": forms.PasswordInput(render_value=True)
        }
    #不允许添加相同用户名
    def clean_username(self):
        username_obj = self.cleaned_data["username"]
        exists = models.admin.objects.filter(username=username_obj)
        if exists:
            raise ValidationError("该用户名已存在")
        return username_obj
    # MD5加密
    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)
    #二次输入的密码也MD5加密
    def clean_confirm_password(self):
        pwd = self.cleaned_data.get("password")
        #对二次输入的密码也加密
        confirm = md5(self.cleaned_data.get("confirm_password"))
        if confirm!=pwd:
            raise ValidationError("两次密码不相同")
        return confirm

##添加管理员
def admin_add(request):
    title = "新建管理员"
    if request.method == "GET":
        form = AdminModelForm()
        return render(request, 'change.html',{'form': form ,"title": title })

    form = AdminModelForm(data=request.POST)
    if form.is_valid():
        #展示所有通过验证的数据
        #print(form.cleaned_data)
        #{'username': 'root', 'password': '123', 'confirm_password': '123'}
        form.save()
        return redirect("/admin/list/")

    return render(request, 'change.html',{'form': form ,"title": title })


#编辑管理员的modelform
class AdminEditModelFrom(bootstrapModelForm):
    #编辑页面密码不可编辑
    password = forms.CharField(disabled=True,label="密码")
    class Meta:
        model = models.admin
        fields = ["username","password"]

#管理员编辑
def admin_edit(request,uid):
    title = "编辑管理员"
    #先获取一个对象，判断是否存在
    admin_obj = models.admin.objects.filter(id=uid).first()
    if not admin_obj:
        return render(request,"404.html",{"error_msg": "404 NOT FOUND!!"})
    if request.method == 'GET':
        #显示原来的值
        form = AdminEditModelFrom(instance=admin_obj)
        return render(request,'change.html',{"title": title,"form": form})

    form = AdminEditModelFrom(data=request.POST, instance=admin_obj)
    if form.is_valid():
        form.save()
        return redirect("/admin/list/")
    return render(request, 'change.html', {"title": title, "form": form})

#管理员删除
def admin_delete(request,uid):
    admin_obj = models.admin.objects.filter(id=uid).first()
    if not admin_obj:
        return render(request,"404.html",{"error_msg": "404 NOT FOUND!!"})
    admin.objects.filter(id=uid).delete()
    return redirect("/admin/list/")


#重置密码modelform
class AdminResetModelForm(bootstrapModelForm):
    confirm_password = forms.CharField(
        label="确认密码",
        widget=forms.PasswordInput(render_value=True)
    )
    class Meta:
        model = models.admin
        fields = ["password","confirm_password"]
        widgets ={
            "password": forms.PasswordInput(render_value=True)
        }
    # MD5加密
    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        md5_pwd = md5(pwd)
        #在数据库校验更改的密码是否与之前一致
        #self.instance.pk，获取当前的id ,查找数据库中该条数据是否有这个密码
        exists = models.admin.objects.filter(id=self.instance.pk,password=md5_pwd).exists()
        if exists:
            raise ValidationError("不允许与之前密码一致")
        return md5_pwd

    #二次输入的密码也MD5加密
    def clean_confirm_password(self):
        pwd = self.cleaned_data.get("password")
        #对二次输入的密码也加密
        confirm = md5(self.cleaned_data.get("confirm_password"))
        if confirm!=pwd:
            raise ValidationError("两次密码不相同")
        return confirm


#管理员重置密码
def admin_reset(request,uid):
    admin_obj = models.admin.objects.filter(id=uid).first()
    if not admin_obj:
        return render(request, "404.html", {"error_msg": "404 NOT FOUND!!"})

    title = "重置密码 - {}".format(admin_obj.username)

    if request.method == "GET":
        form = AdminResetModelForm()
        return render(request, 'change.html', {"title": title,"form": form})

    form = AdminResetModelForm(data=request.POST,instance=admin_obj)
    if form.is_valid():
        form.save()
        return redirect("/admin/list/")
    return render(request,"change.html",{"form": form,"title": title})


##使用Form，自己写字段
class LoginForm(forms.Form):
    username = forms.CharField(
        label="用户名",
        widget = forms.TextInput(attrs={"class": "form-control"}),
        required=True
    )
    password = forms.CharField(
        label="密码",
        widget = forms.PasswordInput(attrs={"class": "form-control"},render_value=True),
        required=True
    )
    code = forms.CharField(
        label="验证码",
        widget = forms.TextInput(attrs={"class": "form-control"}),
        required = True
    )
    ##md5加密
    def clean_password(self):
        pwd = self.cleaned_data.get("password")
        return md5(pwd)


def admin_login(request):
    if request.method == "GET":
        form = LoginForm()
        return render(request,"login.html",{"form": form})

    form = LoginForm(data=request.POST)
    if form.is_valid():
        #表示验证成功，拿到用户名密码
        #form.cleaned_data

        #验证码校验
        #由于数据库中没有code,所有要剔除掉，只保留用户名密码
        user_input_code = form.cleaned_data.pop('code')
        image_code = request.session.get('image_code',"")
        if image_code.upper() != user_input_code.upper():
            form.add_error("code", "验证码错误")
            return render(request, 'login.html', {"form": form})

        # 通过数据库校验
        # 复杂：models.admin.objects.filter(username="xxx",password="xxx").first()
        admin_obj = models.admin.objects.filter(**form.cleaned_data).first()
        if not admin_obj:
            form.add_error("password","用户名密码错误")
            return render(request, 'login.html', {"form": form})
        #输入正确继续操作
        #网站生成随机字符串，写到用户浏览器的cookie中，；再写入session中
        #将info（代表用户的身份信息）自动保存到服务器
        request.session["info"] = {'id': admin_obj.id ,'name': admin_obj.username}

        #重新设置session超时时间
        request.session.set_expiry(60*60*24*7)
        return redirect('/admin/list/')
    return render(request,'login.html',{"form": form})

# 注销
def logout(request):
    #清除会话数据，在存储中删除会话的整条数据
    request.session.clear()
    return redirect('/login/')

from io import BytesIO
#生成图片验证码
def image_code(request):
    #调用pillow函数，生成图片
    img, code = check_code()
    #print(code)

    #将生成的验证码保存到用户session中(以便后续获取验证码进行校验)
    request.session['image_code'] = code
    #设置session60秒超时
    request.session.set_expiry(60)

    #图片内容写入内存中，
    stream = BytesIO()
    img.save(stream,'png')
    #从内存中把原始数据取到
    return HttpResponse(stream.getvalue())


#订单列表的modelform
class orderModelForm(bootstrapModelForm):
    #num = forms.CharField(disabled=True,label="订单号")
    class Meta:
        model = models.order
        #fields = "__all__"
        #fields = {""}
        exclude = ["num","admin"]


def order_list(request):
    order_obj = models.order.objects.all().order_by('id')
    form = orderModelForm
    context={
        "form": form,
        "order_obj": order_obj,
    }
    return render(request,'order_list.html',context)


from django.views.decorators.csrf import csrf_exempt
@csrf_exempt #免除csrf_token认证
## 新建订单，通过ajax请求
def order_add(request):
    form = orderModelForm(data = request.POST)
    if form.is_valid():

        #创建一个num(订单编号)一起保存
        form.instance.num = datetime.now().strftime("%Y%M%d%H%M%S") + str(random.randint(1000,9999))

        #设置一个管理员的信息一起写入数据库 （当前登陆的session中拿到id）
        form.instance.admin_id = request.session["info"]["id"]

        form.save()
        return JsonResponse({"status": True})
       # return HttpResponse(json.dumps({"status": True}))

    return JsonResponse({"status": False, 'error': form.errors})

#订单删除
def order_delete(request):
    nid = request.GET.get("nid")
    exists = models.order.objects.filter(id=nid).exists()
    if not exists:
        return JsonResponse({"status": False,"error": "删除失败,数据不存在"})
    models.order.objects.filter(id=nid).delete()
    return JsonResponse({"status": True})

#编辑展示订单详情
def order_detail(request):
    """ 方式一 ：    根据ID获取订单详情
    nid = request.GET.get("nid")
    //获取到一个对象。
    order_obj = models.order.objects.filter(id=nid).first()
    if not order_obj:
        return JsonResponse({"status": False,"error": "数据不存在"})
    #从数据库获取对象，order_obj
    order_dict = {
        "status": True,
        "data":{
            "title": order_obj.title,
            "price": order_obj.price,
            "status": order_obj.status,
        }
    }
    return JsonResponse(order_dict)
    """
    #方式二
    nid = request.GET.get("nid")
    # 获取到一个对象。
    order_dict = models.order.objects.filter(id=nid).values("title", "price", "status").first()
    if not order_dict:
        return JsonResponse({"status": False, "error": "数据不存在"})

    order_obj = {
        "status": True,
        "data": order_dict
    }
    return JsonResponse(order_obj)

#订单编辑
@csrf_exempt
def order_edit(request):
    nid = request.GET.get("nid")
    order_obj = models.order.objects.filter(id=nid).first()
    if not order_obj:
        return JsonResponse({"status": False, "tips": "数据不存在"})

    form = orderModelForm(data=request.POST,instance = order_obj)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})

    return JsonResponse({"status": False, "error": form.errors})


#数据统计展示页面
def chart_list(request):
    return render(request,"chart_list.html")

#构造柱状图的数据
def chart_bar(request):
    # #数据库中获取数据
    legend = ["销量","业绩"]
    # x_axis = ['衬衫', '羊毛衫', '雪纺衫', '裤子', '高跟鞋', '袜子']
    # series_list = [
    #       {
    #         "name": '销量',
    #         "type": 'bar',
    #         "data": [5, 20, 36, 10, 10, 20]
    #       },
    #       {
    #         "name": '业绩',
    #         "type": 'bar',
    #         "data": [1, 40, 23, 10, 10, 5]
    #       }
    #     ]
    # result = {
    #     "status":True,
    #     "data":{
    #         "legend":legend,
    #         "x_axis":x_axis,
    #         "series_list":series_list
    #     }
    # }
    # return JsonResponse(result)
    objs = models.UserInfo.objects.all()
    return render(request,"chart_list.html",locals())


#构造饼图数据
def chart_pie(request):
    user_obj = models.UserInfo.objects.all()

    db_data_list = [
                { "value": 1048, "name": '公安部' },
                { "value": 735, "name": '交通部' },
                { "value": 580, "name": '运营' },
              ]

    result = {
        "status": True,
        "data": db_data_list,
    }
    return JsonResponse(result)


#上传文件
def upload_list(request):
    if request.method == "GET":
        return render(request,"upload_list.html")

    pic = request.FILES.get('avatar',None)
    #print(pic.name)  #文件名，a1.jpg

    f = open(pic.name,mode='wb')
    for chunk in pic.chunks():
        f.write(chunk)
    f.close()

    return HttpResponse(",,,")


## 文件上传From,
class UpForm(forms.Form):
    name = forms.CharField(label="姓名")
    age = forms.IntegerField(label="年龄")
    img = forms.FileField(label="头像")

    # 可以局部添加/移除样式
    bootstrap_exclude_fields = ['img']
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        # 循环找到所有插件，添加class
        for name,field in self.fields.items():
            if name in self.bootstrap_exclude_fields:
                continue
            if field.widget.attrs :
                field.widget.attrs["class"] = "form-control"
            else:
                field.widget.attrs = {
                    "class": "form-control"
                }

# 文件上传，Form实现组合数据上传
def upload_form(request):
    title = "FORM文件上传"
    if request.method == "GET":
        form = UpForm()
        return render(request,'upload_form.html',{"form": form,"title":title})

    form = UpForm(data=request.POST,files=request.FILES)
    if form.is_valid():
        #1.读取图片内容，写入一个文件夹，并获取路径，
        img_obj = form.cleaned_data.get("img")

        #绝对路径
        #media_path = os.path.join(settings.MEDIA_ROOT,img_obj.name)

        #相对路径
        media_path = os.path.join("media", img_obj.name)

        #file_path = "app01/static/img/{}".format(img_obj.name)
        #file_path = os.path.join("app01",media_path)

        f = open(media_path,mode="wb")
        for chunk in img_obj.chunks():
            f.write(chunk)
        f.close()

        #2.将图片路径存入数据库
        models.Boss.objects.create(
            name=form.cleaned_data['name'],
            age = form.cleaned_data['age'],
            img = media_path
        )

        return HttpResponse("---")
    return render(request,'upload_form.html',{"form": form,"title":title})


#文件上传的ModelForm
class UpModelForm(bootstrapModelForm):
    bootstrap_exclude_fields = ['img']
    class Meta:
        model = models.City
        fields = "__all__"


# 文件上传，modelForm实现
def upload_modelform(request):
    title = "ModelForm上传"
    if request.method == "GET":
        form = UpModelForm()
        return render(request,'upload_form.html',{"form":form,"title":title})

    form = UpModelForm(data=request.POST,files=request.FILES)
    if form.is_valid():
        #写入数据库，自动将文件保存起来，将路径保存到数据库
        form.save()
        return HttpResponse("yes")
    return render(request, 'upload_form.html', {"form": form, "title": title})




def city_list(request):
    city_obj = models.City.objects.all()
    return render(request,"city_list.html",{"city_obj":city_obj})

def city_add(request):
    title = "新建城市"
    if request.method == "GET":
        form = UpModelForm()
        return render(request, 'upload_form.html', {"form": form, "title": title})

    form = UpModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        # 写入数据库，自动将文件保存起来，将路径保存到数据库
        form.save()
        return redirect("/city/list/")
    return render(request, 'upload_form.html', {"form": form, "title": title})