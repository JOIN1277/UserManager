# author:klx
# time:2022/6/4 -{TIME}
# function:
#
from django.http import HttpResponse
from django.shortcuts import render,redirect
from app01.models import Department,UserInfo,Number,admin,Boss,City
from django import forms
from app01 import models
from django.core.validators import RegexValidator,ValidationError
from app01.utils.bootstrap import bootstrapModelForm
from app01.utils.encrypt import md5
from app01.utils.code import check_code
from django.http import JsonResponse
import random
import os
from datetime import datetime
from django.conf import settings
from app01.utils.form import UserModelFrom,PrettyModelFrom,AdminModelForm,AdminEditModelFrom,AdminResetModelForm,orderModelForm,UpForm,UpModelForm,PrettyEditModelFrom
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
# Create your views here.
def depart_show(request):
    #获取部门表中的数据
    department = Department.objects.all()
    return render(request,"depart_list.html",{'department':department})

def depart_add(request):
    #新增部门数据
    if request.method =="GET":
        return render(request,"depart_add.html")
    title = request.POST.get("title")
    Department.objects.create(title=title)
    return redirect("/depart/list/")

def depart_delete(request):
    #删除部门
    #通过uid找到当前部门信息
    uid = request.GET.get('uid')
    #删除
    Department.objects.filter(id=uid).delete()
    #跳转回列表
    return redirect("/depart/list/")

def depart_edit(request,uid):
    if request.method == "GET":
        departmentById = Department.objects.filter(id=uid).first()
        return render(request,"depart_edit.html",{'departmentById':departmentById})
    #根据id找到当前数据库中的数据
    title_Edit = request.POST.get("title")
    Department.objects.filter(id=uid).update(title=title_Edit)
    return redirect("/depart/list/")

#批量上传excel文件
def depart_multi(request):

    # 1.获取文件的对象
    exc_obj = request.FILES.get("exc")

    # 2.对象传递给openpyxl,由openpyxl读取文件内容
    wb = load_workbook(exc_obj)
    sheet = wb.worksheets[0] #excel的第一个sheet

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):  # 最小行2，从第2行开始
        text = row[0].value
        #print(text)
        #判断
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect("/depart/list/")
